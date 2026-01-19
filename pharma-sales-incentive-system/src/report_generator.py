"""
Pharma Sales Incentive System - Rapor Oluşturma Modülü
Excel ve özet raporları oluşturma
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
from datetime import datetime

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ReportGenerator:
    """Rapor oluşturma sınıfı"""
    
    def __init__(self):
        # Stil tanımları
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        self.success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.danger_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _apply_header_style(self, ws, row: int, start_col: int, end_col: int):
        """Başlık stilini uygula"""
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
    
    def _apply_data_style(self, ws, start_row: int, end_row: int, start_col: int, end_col: int):
        """Veri stilini uygula"""
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _get_score_fill(self, score: float, threshold: float = 85) -> PatternFill:
        """Skora göre renk belirle"""
        if score >= threshold:
            return self.success_fill
        elif score >= threshold * 0.9:
            return self.warning_fill
        else:
            return self.danger_fill
    
    def generate_st_scorecard(
        self,
        st_scores: List[Any],
        output_path: str,
        quarter: int,
        year: int
    ) -> str:
        """ST Scorecard raporu oluştur"""
        wb = Workbook()
        
        # Özet sayfası
        ws_summary = wb.active
        ws_summary.title = "Özet"
        
        # Başlık
        ws_summary['A1'] = f"ST Performans Scorecard - Q{quarter} {year}"
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.merge_cells('A1:H1')
        
        # Başlıklar
        headers = ['ST Adı', 'BM', 'Bölge', 'Satış Skoru', 'Rota Skoru', 
                   'Sipariş Skoru', 'Toplam Skor', 'Prim Hakkı']
        
        for col, header in enumerate(headers, 1):
            ws_summary.cell(row=3, column=col, value=header)
        self._apply_header_style(ws_summary, 3, 1, len(headers))
        
        # Veriler
        for idx, st in enumerate(st_scores, 4):
            ws_summary.cell(row=idx, column=1, value=st.st_name)
            ws_summary.cell(row=idx, column=2, value=st.bm_name)
            ws_summary.cell(row=idx, column=3, value=st.region)
            ws_summary.cell(row=idx, column=4, value=round(st.kpi_scores.get('sales_volume', 0), 2))
            ws_summary.cell(row=idx, column=5, value=round(st.kpi_scores.get('route_compliance', 0), 2))
            ws_summary.cell(row=idx, column=6, value=round(st.kpi_scores.get('order_success', 0), 2))
            ws_summary.cell(row=idx, column=7, value=round(st.total_score, 2))
            ws_summary.cell(row=idx, column=8, value="EVET" if st.eligible_for_bonus else "HAYIR")
            
            # Renklendirme
            score_cell = ws_summary.cell(row=idx, column=7)
            score_cell.fill = self._get_score_fill(st.total_score)
            
            bonus_cell = ws_summary.cell(row=idx, column=8)
            bonus_cell.fill = self.success_fill if st.eligible_for_bonus else self.danger_fill
        
        self._apply_data_style(ws_summary, 4, len(st_scores) + 3, 1, len(headers))
        
        # Kolon genişlikleri
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20
        ws_summary.column_dimensions['C'].width = 20
        for col in 'DEFGH':
            ws_summary.column_dimensions[col].width = 15
        
        # Detay sayfası
        ws_detail = wb.create_sheet("Ürün Detay")
        ws_detail['A1'] = "Ürün Bazlı Performans Detayı"
        ws_detail['A1'].font = Font(bold=True, size=14)
        
        detail_headers = ['ST Adı', 'Ürün Grubu', 'Hedef', 'Gerçekleşen', 
                         'Oran (%)', 'Ağırlık', 'Puan']
        
        for col, header in enumerate(detail_headers, 1):
            ws_detail.cell(row=3, column=col, value=header)
        self._apply_header_style(ws_detail, 3, 1, len(detail_headers))
        
        row = 4
        for st in st_scores:
            for product, scores in st.product_scores.items():
                ws_detail.cell(row=row, column=1, value=st.st_name)
                ws_detail.cell(row=row, column=2, value=product)
                ws_detail.cell(row=row, column=3, value=scores.get('target', 0))
                ws_detail.cell(row=row, column=4, value=scores.get('actual', 0))
                ws_detail.cell(row=row, column=5, value=round(scores.get('rate', 0) * 100, 1))
                ws_detail.cell(row=row, column=6, value=scores.get('weight', 0))
                ws_detail.cell(row=row, column=7, value=round(scores.get('score', 0), 2))
                row += 1
        
        # Kaydet
        wb.save(output_path)
        logger.info(f"ST Scorecard raporu oluşturuldu: {output_path}")
        return output_path
    
    def generate_bm_report(
        self,
        bm_scores: List[Dict],
        output_path: str,
        quarter: int,
        year: int
    ) -> str:
        """BM Özet raporu oluştur"""
        wb = Workbook()
        ws = wb.active
        ws.title = "BM Özet"
        
        # Başlık
        ws['A1'] = f"Bölge Müdürü Performans Özeti - Q{quarter} {year}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        # Başlıklar
        headers = ['BM Adı', 'ST Sayısı', 'Ortalama Skor', 'Prim Hak Eden', 
                   'Prim Oranı (%)', 'Değerlendirme']
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
        self._apply_header_style(ws, 3, 1, len(headers))
        
        # Veriler
        for idx, bm in enumerate(bm_scores, 4):
            ws.cell(row=idx, column=1, value=bm['bm_name'])
            ws.cell(row=idx, column=2, value=bm['st_count'])
            ws.cell(row=idx, column=3, value=round(bm['avg_score'], 2))
            ws.cell(row=idx, column=4, value=bm['eligible_count'])
            ws.cell(row=idx, column=5, value=round(bm['eligible_rate'], 1))
            
            # Değerlendirme
            if bm['eligible_rate'] >= 80:
                evaluation = "Mükemmel"
            elif bm['eligible_rate'] >= 60:
                evaluation = "İyi"
            elif bm['eligible_rate'] >= 40:
                evaluation = "Orta"
            else:
                evaluation = "Geliştirilmeli"
            
            ws.cell(row=idx, column=6, value=evaluation)
        
        self._apply_data_style(ws, 4, len(bm_scores) + 3, 1, len(headers))
        
        # Kolon genişlikleri
        ws.column_dimensions['A'].width = 25
        for col in 'BCDEF':
            ws.column_dimensions[col].width = 15
        
        wb.save(output_path)
        logger.info(f"BM raporu oluşturuldu: {output_path}")
        return output_path
    
    def generate_product_analysis(
        self,
        st_scores: List[Any],
        output_path: str,
        quarter: int,
        year: int
    ) -> str:
        """Ürün grubu analiz raporu"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Ürün Analizi"
        
        # Başlık
        ws['A1'] = f"Ürün Grubu Performans Analizi - Q{quarter} {year}"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Ürün bazlı aggregasyon
        product_totals = {}
        
        for st in st_scores:
            for product, scores in st.product_scores.items():
                if product not in product_totals:
                    product_totals[product] = {
                        'total_target': 0,
                        'total_actual': 0,
                        'st_count': 0,
                    }
                product_totals[product]['total_target'] += scores.get('target', 0)
                product_totals[product]['total_actual'] += scores.get('actual', 0)
                product_totals[product]['st_count'] += 1
        
        # Başlıklar
        headers = ['Ürün Grubu', 'Toplam Hedef', 'Toplam Gerçekleşen', 
                   'Oran (%)', 'ST Sayısı']
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
        self._apply_header_style(ws, 3, 1, len(headers))
        
        # Veriler
        row = 4
        for product, totals in sorted(product_totals.items()):
            rate = (totals['total_actual'] / totals['total_target'] * 100 
                   if totals['total_target'] > 0 else 0)
            
            ws.cell(row=row, column=1, value=product)
            ws.cell(row=row, column=2, value=totals['total_target'])
            ws.cell(row=row, column=3, value=totals['total_actual'])
            ws.cell(row=row, column=4, value=round(rate, 1))
            ws.cell(row=row, column=5, value=totals['st_count'])
            
            # Renklendirme
            rate_cell = ws.cell(row=row, column=4)
            if rate >= 100:
                rate_cell.fill = self.success_fill
            elif rate >= 85:
                rate_cell.fill = self.warning_fill
            else:
                rate_cell.fill = self.danger_fill
            
            row += 1
        
        self._apply_data_style(ws, 4, row - 1, 1, len(headers))
        
        # Kolon genişlikleri
        ws.column_dimensions['A'].width = 30
        for col in 'BCDE':
            ws.column_dimensions[col].width = 18
        
        wb.save(output_path)
        logger.info(f"Ürün analiz raporu oluşturuldu: {output_path}")
        return output_path
    
    def generate_bonus_report(
        self,
        st_scores: List[Any],
        output_path: str,
        quarter: int,
        year: int
    ) -> str:
        """Prim hak eden ST raporu"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Prim Listesi"
        
        # Başlık
        ws['A1'] = f"Prim Hak Kazanan Satış Temsilcileri - Q{quarter} {year}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        # Özet bilgi
        eligible = [s for s in st_scores if s.eligible_for_bonus]
        ws['A2'] = f"Toplam: {len(eligible)} / {len(st_scores)} ST prim hakkı kazandı"
        ws['A2'].font = Font(italic=True)
        
        # Başlıklar
        headers = ['Sıra', 'ST Adı', 'BM', 'Bölge', 'Toplam Skor']
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=header)
        self._apply_header_style(ws, 4, 1, len(headers))
        
        # Veriler (skoruna göre sıralı)
        eligible_sorted = sorted(eligible, key=lambda x: x.total_score, reverse=True)
        
        for idx, st in enumerate(eligible_sorted, 1):
            row = idx + 4
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=st.st_name)
            ws.cell(row=row, column=3, value=st.bm_name)
            ws.cell(row=row, column=4, value=st.region)
            ws.cell(row=row, column=5, value=round(st.total_score, 2))
            
            # Tüm satırı yeşil yap
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = self.success_fill
        
        self._apply_data_style(ws, 5, len(eligible_sorted) + 4, 1, len(headers))
        
        # Kolon genişlikleri
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        
        wb.save(output_path)
        logger.info(f"Prim raporu oluşturuldu: {output_path}")
        return output_path
    
    def generate_all_reports(
        self,
        st_scores: List[Any],
        bm_scores: List[Dict],
        output_dir: str,
        quarter: int,
        year: int
    ) -> Dict[str, str]:
        """Tüm raporları oluştur"""
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        reports = {}
        
        # ST Scorecard
        reports['st_scorecard'] = self.generate_st_scorecard(
            st_scores,
            str(output_path / f"ST_Scorecard_Q{quarter}_{year}_{timestamp}.xlsx"),
            quarter, year
        )
        
        # BM Raporu
        reports['bm_report'] = self.generate_bm_report(
            bm_scores,
            str(output_path / f"BM_Ozet_Q{quarter}_{year}_{timestamp}.xlsx"),
            quarter, year
        )
        
        # Ürün Analizi
        reports['product_analysis'] = self.generate_product_analysis(
            st_scores,
            str(output_path / f"Urun_Analiz_Q{quarter}_{year}_{timestamp}.xlsx"),
            quarter, year
        )
        
        # Prim Listesi
        reports['bonus_report'] = self.generate_bonus_report(
            st_scores,
            str(output_path / f"Prim_Listesi_Q{quarter}_{year}_{timestamp}.xlsx"),
            quarter, year
        )
        
        logger.info(f"Tüm raporlar oluşturuldu: {output_dir}")
        return reports
